# Librería para trabajar con fechas y horas - Usada para: generar número único de factura y obtener fecha/hora actual automáticamente
from datetime import datetime as dt
# Define el tamaño de página para el PDF - Usada para: establecer que el PDF sea tamaño carta (8.5" x 11")
from reportlab.lib.pagesizes import letter
# Librería de colores para PDFs 
from reportlab.lib import colors
# Elementos para construir el layout del PDF - Usados para: crear la estructura del PDF (documento base, tablas de info/productos/totales, título, espacios entre secciones)
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
# Estilos de texto para el PDF - Usados para: obtener estilos predefinidos y crear el estilo personalizado del título azul centrado de 24pt
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# Unidad de medida para dimensiones en el PDF - Usada para: definir anchos de columnas (2*inch, 5*inch) y espacios verticales (0.3*inch) de forma legible
from reportlab.lib.units import inch
# Librería para trabajar con archivos Excel - Usada para: abrir archivos Excel existentes con load_workbook() y verificar si el archivo de registro existe
import openpyxl
# Clase para crear nuevos archivos Excel - Usada para: crear el archivo "Registro_Ventas.xlsx" desde cero la primera vez que se ejecuta el programa
from openpyxl import Workbook
# Módulo del sistema operativo - Usado para: verificar con os.path.exists() si los archivos (Excel y PDF) ya existen antes de crearlos o abrirlos
import os

#Creamos un diccionario dentro de otro diccionario con los productos disponibles, accedemos mediante el codigo del producto
productos = {
     1: {"nombre": "Unicornio mediano 1", "precio": 5700},
     2: {"nombre": "Unicornio Mediano 2", "precio": 6800},    
     3: {"nombre": "Unicornio Mediano 3", "precio": 8000},
     4: {"nombre": "Elefante Trompeta #1", "precio": 5750},
     5: {"nombre": "Elefante Trompeta #2", "precio": 6650},
     6: {"nombre": "Corazon Realista", "precio": 9000},
     7: {"nombre": "Perro Globo Mini", "precio": 3000},
     8: {"nombre": "Perro Globo Mediano", "precio": 9000},
     9: {"nombre": "Perro Globo Super", "precio": 15000},
    10: {"nombre": "Groot Canasta", "precio": 7000},    
    11: {"nombre": "Groot Tronco", "precio": 6750},
    12: {"nombre": "Dinosaurio", "precio": 7000},
    13: {"nombre": "Lagarto", "precio": 8000},
    14: {"nombre": "Zorro Sentado", "precio": 5350},
    15: {"nombre": "Zorro Acostado", "precio": 7450},
    16: {"nombre": "Mafalda", "precio": 13500},
    17: {"nombre": "Yogas", "precio": 10000},
    18: {"nombre": "Mano Corazon", "precio": 6750},
    19: {"nombre": "Astronauta DJ", "precio": 10000},
    20: {"nombre": "Astronautas", "precio": 7350},
    21: {"nombre": "Osos Escandalosos", "precio": 6000},
    22: {"nombre": "Capibara Morral", "precio": 6750},
    23: {"nombre": "Capibara Liso", "precio": 6750},
    24: {"nombre": "Unicornio Liso Grande ", "precio": 9750},
    25: {"nombre": "Unicornio Alas Grande ", "precio": 10000},
    26: {"nombre": "Pokemon Pikachu", "precio": 6000},
    27: {"nombre": "Perro Gafas", "precio": 5450},
    28: {"nombre": "Creaneo Super", "precio": 9750},
    29: {"nombre": "Creaneo Mini ", "precio": 3000},
    30: {"nombre": "Careneo Decorativo Colgante", "precio": 8750}
}

#Creamos una funcion para mostrar los productos disponibles
def mostrar_productos():
    print("\n" + "="*60)      
    print("------------------ CATÁLOGO DE PRODUCTOS -------------------")    
    print("="*60)    
    # Recorremos el diccionario e imprimimos cada producto
    #Creamos un for con dos variables, una para el codigo y otra para la infomación del producto
    #Utilizamos items() para obtener ambos valores del diccionario
    #Accedemos al nombre y precio mediante las claves del diccionario interno con info['nombre'] y info['precio']
    for codigo, info in productos.items():
        print(f"{codigo}. Producto: {info['nombre']} - Precio: ${info['precio']}")
    print("="*60)

#Funcion para obtener la fecha y hora actual en formato dd/mm/yyyy hh:mm:ss
def fecha_actual():
    #el return devuelve la fecha y hora actual cada vez que se llama a la funcion
    return dt.now().strftime("%d/%m/%Y %H:%M:%S")

#Funcion para generar un numero unico de factura basado en la fecha y hora actual
def generar_numero_factura():
    #utiilizamos datetime para obtener la fecha el .now() y luego formateamos con strftime para obtener un string unico
    return dt.now().strftime("%Y%m%d%H%M%S")

#Funcion para calcular el descuento del 5% para empresas
def calcular_descuento(subtotal):
    descuento = 0
    descuento=subtotal*0.05
    total= subtotal - descuento
    # Retornamos el descuento y el total para usarlos despues
    return descuento, total

#Función para recopilar los datos de facturación
def datos_factura():
    print("="*60)
    print("----------- SISTEMA DE FACTURACIÓN - CERÁMICAS  -----------")
    print("="*60)

    #Le asignamos a una variable la funcion de generar numero de factura
    global numero_unico_factura
    numero_unico_factura = generar_numero_factura()
    print(f"Número de factura: {numero_unico_factura}")

    print("\n---- DATOS DE FACTURACIÓN ----")

    #Le asignamos a una variable la funcion de fecha actual
    global fecha
    fecha = fecha_actual()
    print(f"\nFecha y hora: {fecha}")

    #creamos las variables para almacenar los datos del cliente y del facturador y validamos que no esten vacias
    global nombre_facturador
    global nombre_cliente
    while True:
        nombre_facturador = input("\nNombre de quien factura: ")
        nombre_cliente = input("\nNombre del cliente o empresa: ")
        if nombre_facturador=="" or nombre_cliente=="":
            print("\nIngrese los nombres correspondientes:")
            continue
        else:
            break

    #Creamos un while para validar la entrada del tipo de cliente
    while True:
        tipo_cliente = input("\n¿Es empresa proveedora? (si/no): ").lower()
        if tipo_cliente in ['si', 'no']:
            break
        else:
            print(f"{tipo_cliente}, No es una entrada inválida. Por favor ingrese 'si' o 'no'.")
    #Asignamos el tipo de cliente dependiendo de la respuesta 
    if tipo_cliente == 'si':
        tipo = "Empresa (Proveedor)"
    else:
        tipo = "Cliente Regular"
    # Booleano para saber si es empresa o no
    global es_empresa
    es_empresa = tipo_cliente == 'si'
    # Diccionario para almacenar los datos del cliente
    #Este diccionario se usa para mostrar los datos en el resumen de la compra
    global datos_cliente
    datos_cliente = {
        'cliente': nombre_cliente, 
        'tipo': tipo,  
        'facturador': nombre_facturador,  
        'fecha': fecha  
    }
    # Retornamos los datos recopilados
    return datos_cliente, es_empresa, numero_unico_factura, fecha, nombre_facturador, nombre_cliente, tipo, tipo_cliente


#creamos una funcion para guardar los datos de la venta en un archivo excel
def guardar_en_excel(datos_cliente, items_compra, total, numero_unico_factura):
    #creamos una variable con el nombre del archivo excel y le asiganamos el nombre del Excel con la extension .xlsx
    archivo_excel = "Registro_Ventas_Ceramicas.xlsx"  
    # try: intenta ejecutar este código
    # except: si hay un error, ejecuta esto en lugar de cerrar el programa
    try:
        # os.path es un modulo del paquete os y el .exists() es un metodo que verifica si un archivo existe
        # archivo_excel es el ARGUMENTO que pasamos y devuelve True si existe, False si no existe
        if os.path.exists(archivo_excel):
            # Si el archivo existe, entramos aquí, solo se ejecuta si es True
            # openpyxl.load_workbook es una función que abre un Excel y pasamos el nombre del archivo como argumento
            #y le asigna el resultado a la variable libro para trabajar con el Excel
            libro = openpyxl.load_workbook(archivo_excel)
            # .active obtiene la hoja activa del libro abierto
            hoja = libro.active  
            # hoja es un objeto Worksheet(Un Worksheet es un objeto que representa una hoja de un archivo Excel.)
        # Si el archivo NO existe, entramos aquí
        else:
            # Workbook() crea un nuevo libro de Excel vacío 
            libro = Workbook()
            # libro es un nuevo objeto Workbook(Un Workbook es un objeto que representa un archivo completo de Excel (el archivo .xlsx).)
            # .active obtiene la hoja activa del nuevo libro
            hoja = libro.active
            # Cambiamos el nombre de la hoja activa a "Ventas"
            hoja.title = "Ventas"  
            # Creamos una LISTA con los nombres de las columnas
            # Cada string es el encabezado de una columna
            encabezados = ["Factura N°", "Fecha", "Cliente/Empresa", "Tipo", "Facturado por", 
                          "Producto", "Cantidad", "Precio Unit.", "Total Item", "Total Factura"]
            # .append() es un metodo que agrega una fila al Excel y pasamos como argumento la lista, esto crea la primera fila con los títulos
            hoja.append(encabezados)  
        
        # Ahora agregamos los datos de la venta al Excel
        # Recorremos cada item comprado para agregar una fila por cada producto
        # item es la variable que toma cada elemento y items_compra es la lista de productos comprados
        for item in items_compra:
            # Creamos una LISTA con todos los datos de una fila, cada elemento es una CELDA, ['clave'] accede al valor en el diccionario
            fila = [
                numero_unico_factura,  # Columna A: número de factura
                datos_cliente['fecha'],  # Columna B: fecha
                datos_cliente['cliente'],  # Columna C: nombre cliente
                datos_cliente['tipo'],  # Columna D: tipo (empresa o cliente)
                datos_cliente['facturador'],  # Columna E: quien factura
                item['nombre'],  # Columna F: nombre del producto
                item['cantidad'],  # Columna G: cantidad
                item['precio'],  # Columna H: precio unitario
                item['total'],  # Columna I: total del item
                total  # Columna J: total de la factura
            ]
            # .append() agrega la fila al final del Excel
            hoja.append(fila)  
            # Cada vuelta del for agrega una nueva fila
        
        # Después del for, todas las filas están agregadas en memoria, pero aún NO están guardadas en el archivo físico
        # Asi que usamos .save() para guardar los cambios y en los argumentos le damos el nombre del archivo
        libro.save(archivo_excel)
        # Si todo sale bien, mostraremos mensaje de éxito
        print(f"Venta registrada en Excel: {archivo_excel}")
        
    # PermissionError es un tipo de error que ocurre cuando(No hay permisos para acceder al archivo, el esra abierto en otro programa, etc):
    except PermissionError:
        print("\n  ERROR: No se pudo guardar en Excel")
        print("El archivo '{archivo_excel}' está abierto en otro programa.")
        print("Por favor CIERRE el archivo Excel e intente nuevamente.")
        
    # EXCEPT sin especificar tipo captura cualquier otro error y el as e guarda la información del error en la varaible e
    except Exception as e:
        print(f"\nError al guardar en Excel: {e}\n")


def crear_factura_pdf(datos_cliente, items_compra, subtotal, descuento, total, numero_unico_factura):
    #Creamos una variable con el nombre del archivo pdf usando f-string para incluir el numero de factura y el nombre del cliente y con el replace para eliminar espacios y poner guiones bajos
    nombre_archivo = f"Factura_{numero_unico_factura}_{datos_cliente['cliente'].replace(' ', '_')}.pdf"
    #Creamos una variable pdf a la cual le asignamos el objeto SimpleDocTemplate(sirve para crear documentos PDF) con el nombre del archivo y el tamaño de pagina letter
    pdf = SimpleDocTemplate(nombre_archivo, pagesize=letter)
    #Creamos una lista vacía llamada elementos para almacenar los componentes del PDF
    elementos = []
    #Creamos una variable estilos a la cual le asignamos el resultado de la función getSampleStyleSheet() para obtener estilos predefinidos
    estilos = getSampleStyleSheet()
    #Creamos la variable estilo_titulo para definir un estilo personalizado para el título del PDF con ParagraphStyle() el cual le pasa varios parámetros para definir el formato
    estilo_titulo = ParagraphStyle(
        # Nombre del estilo  
        'CustomTitle',  
        # paraent= indica que este estilo HEREDA características de otro estilo y le asignamos el estilo Heading1 predefinido
        parent=estilos['Heading1'],  
        #el fontSize define el tamaño de la letra del título         
        fontSize=24,  
        #Aqui elegimos el color del texto usando un código hexadecimal para un azul
        textColor=colors.HexColor('#1E90FF'),  
        #El spaceBefore define el espacio que va a tener entre el titulo y la tabla de información
        spaceAfter=15, 
        #Alineacion de texto del título el 1 es hacia el centro
        alignment=1  
    )
    #Creamos la variable titulo para crear el objeto Paragraph con el texto y el estilo definido arriba, el Paragraph crea un párrafo de texto formateado
    titulo = Paragraph("FACTURA DE VENTA", estilo_titulo)
    # Agregamos el título a la lista de elementos del PDF con .append()
    elementos.append(titulo)  
    #Agregamos un espacio vertical entre el título y la siguiente sección, con el spacer de 0.3 pulgadas
    #En los argumentos (1, 0.3*inch) definimos el ancho y la altura del espacio, inch es una constante que vale 72 puntos
    elementos.append(Spacer(1, 0.3*inch))  
    #Creamos la variable info_general para almacenar la información general de la factura en formato de tabla, creamos mas listas adentro para cada fila
    info_general = [
        #Cada fila es una lista con 2 elementos (2 columnas)
        ['Factura N°:', numero_unico_factura],  
        ['Fecha:', datos_cliente['fecha']], 
        ['Cliente/Empresa:', datos_cliente['cliente']], 
        ['Tipo:', datos_cliente['tipo']], 
        ['Facturado por:', datos_cliente['facturador']]  
    ]
    #Creamos la variable tabla_info para crear el objeto Table con los datos de info_general y definimos los anchos de las columnas con colWidths
    tabla_info = Table(info_general, colWidths=[2*inch, 5*inch])
    #Creamos los estilos para la tabla de información con .setStyle() aplica los estilos visuales a la tabl a y TableStyle() define el estilo
    #([ ]) = recibe una lista de tuplas, cada tupla es una instrucción de estilo, formato: ('COMANDO', celda_inicio, celda_fin, parámetros)
    tabla_info.setStyle(TableStyle([
        # Primera tupla de estilo:el background de la primera columna es para poner el fondo azul oscuro, de la primera celda (0,0) a la última (0,-1)
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#4169E1')), 
        # Segunda tupla de estilo: el color del texto de la primera columna es blanco humo (casi blanco)
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),            
        # Tercera tupla de estilo: color de fondo de la segunda columna azul muy claro
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#E6F2FF')), 
        # Cuarta tupla de estilo: color del texto de la segunda
        ('TEXTCOLOR', (1, 0), (1, -1), colors.black),                  
        # Quinta tupla de estilo: alineación del texto en TODAS las celdas a la izquierda
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        # Sexta tupla de estilo: tipo de letra de la primera columna en negrita
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        # Séptima tupla de estilo: tamaño de letra en TODAS las celdas
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        # Octava tupla de estilo: espacio interno inferior de cada celda
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        # Novena tupla de estilo: agregar bordes a TODAS las celdas
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#122A6B"))
    ]))
    # Agregamos la tabla de información a la lista de elementos del PDF
    elementos.append(tabla_info)
    # Agregamos un espacio vertical entre la tabla de info y la tabla de productos
    elementos.append(Spacer(1, 0.4*inch))
    #Creamos la variable encabezados para almacenar los títulos de las columnas de la tabla de productos
    encabezados = [['N°', 'Producto', 'Cantidad', 'Precio Unit.', 'Total']]
    # Recorremos cada item comprado para agregar una fila por cada producto
    for item in items_compra:
        # Creamos la variable fila para almacenar los datos de cada producto en una lista
        fila = [
            # Convertimos el número a string para la tabla
            str(item['numero']),  
            item['nombre'],  
            str(item['cantidad']), 
            #Utilizamos f-string para formatear el precio y el total con comas como separadores de miles
            f"${item['precio']:,}", 
            f"${item['total']:,}" 
        ]
        # Agregamos la fila a la lista de encabezados con .append()
        encabezados.append(fila)  
    #creamos una variable tabla_productos para crear el objeto Table con los datos de encabezados y definimos los anchos de las columnas con colWidths
    tabla_productos = Table(encabezados, colWidths=[0.5*inch, 2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
    #Creamos los estilos para la tabla de productos con .setStyle() aplica los estilos visuales a la tabla y TableStyle() define el estilo
    tabla_productos.setStyle(TableStyle([
        # Primera tupla de estilo: fondo azul oscuro para la fila 0 (encabezados)
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4169E1')),
        # Segunda tupla de estilo: texto blanco para la fila 0 (encabezados)
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        # Tercera tupla de estilo: Todas las filas estaran centradas
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        # Cuarta tupla de estilo: tipo de letra en negrita para la fila 0 (encabezados)
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        # Quinta tupla de estilo: tamaño de letra 12 para la fila 0 (encabezados)
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        # Sexta tupla de estilo: espacio interno inferior en la fila 0 (encabezados)
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        # Séptima tupla de estilo: fondo azul muy claro para las filas de datos (1 a -1)
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E6F2FF')),
        # Octava tupla de estilo:agregar bordes a TODAS las celdas
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#122A6B'))
    ]))
    # Agregamos la tabla de productos a la lista de elementos del PDF
    elementos.append(tabla_productos)
    # Agregamos un espacio vertical entre la tabla de productos y la sección de totales
    elementos.append(Spacer(1, 0.3*inch))
    #Validamos atraves de un condicional si hay descuento o no
    if descuento > 0:
        #Si hay descuento mostraremos en el pdf 3 filas: subtotal, descuento y total, creamos listas dentro de otra lista para cada fila     
        datos_totales = [
            ['Subtotal:', f"${subtotal:,}"],
            ['Descuento 5% (Empresa):', f"-${descuento:,}"],  
            ['TOTAL A PAGAR:', f"${total:,}"]
        ]
    else:
        #De lo contrario si no hay descuento mostraremos solo el total a pagar, creamos una lista dentro de otra lista para la fila
        datos_totales = [
            ['TOTAL A PAGAR:', f"${total:,}"]
            # Solo el total (que es igual al subtotal)
        ]
    # Creamos la variable tabla_totales para crear el objeto Table con los datos de datos_totales y definimos los anchos de las columnas con colWidths
    tabla_totales = Table(datos_totales, colWidths=[4*inch, 2*inch])
    # Creamos los estilos para la tabla de totales con .setStyle() aplica los estilos visuales a la tabla y TableStyle() define el estilo
    tabla_totales.setStyle(TableStyle([
        # Alineación del texto en TODAS las celdas a la derecha
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        #La ultima fila en negrita        
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        #Le ponemos negrita a las filas de si hay descuento y subtotal
        ('FONTNAME', (0, -3), (-1, -2), 'Helvetica-Bold'),
        # Tamaño de letra 14 para la última fila        
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        # Color azul para el texto de la última fila
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1E90FF')),
        # Línea gruesa arriba de la última fila
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1A5484')),
        # Espacio interno superior en la última fila        
        ('TOPPADDING', (0, -1), (-1, -1), 15),

    ]))
    # Agregamos la tabla de totales a la lista de elementos del PDF
    elementos.append(tabla_totales)
    # Finalmente, construimos el PDF con .build() y le pasamos la lista de elementos
    pdf.build(elementos)
    # Mostramos mensaje de éxito con el nombre del archivo generado 
    print(f"\nFactura PDF generada: {nombre_archivo}")


#Función principal de facturación
def facturacion():
    datos_factura()
    subtotal=0
    # Lista para almacenar los items de la compra
    items_compra = []
    print("\n SELECCIÓN DE PRODUCTOS:")
    print("(Ingrese 0 para finalizar la compra)")
    # Bucle principal para agregar productos
    while True:
        mostrar_productos()
        # Bucle para validar el número de producto
        while True:
        #try y except para manejar errores de entrada
            try:
                #creamos una variable para almacenar el numero del producto
                numero_producto = int(input("\nIngrese el número del producto (1-30, 0 para salir): "))
                #validamos que el numero este en el rango de productos disponibles
                if numero_producto == 0 or numero_producto in productos:
                    #si es 0 o un numero que esta en el diccionario productos, salimos del bucle
                    break
                else:
                    print(f"{numero_producto}, es un número inválido. Intente de nuevo.")
            except ValueError:
                print("Entrada inválida. Por favor ingrese un número válido.")
        # Si ingresa 0, verificamos si hay productos agregados
        if numero_producto == 0:
            #Con el len verificamos la longitud de la lista items_compra
            if len(items_compra) == 0:
                print("\n No has agregado ningún producto. Debes agregar al menos un producto.")
                # con el continue volvemos al inicio del bucle principal
                continue
            else:
                # Si hay productos, salimos del bucle principal
                break
        # Bucle para validar la cantidad
        while True:
            try:
                cantidad = int(input("Ingrese la cantidad deseada: "))
                if cantidad > 0:
                    break
                else:
                    print(" La cantidad debe ser mayor a 0.")
            except ValueError:
                print("Ingresaste una palabra o un valor que no es valido, por favor ingrese un número válido.")
        #obtenemos la info del producto seleccionado mediante el codigo el cual se ingresa en la variable de numero_producto
        producto_info = productos[numero_producto]
        #calculamos el total del item, donde adcedemos al precio del producto mediante el diccionario productos
        total_item = producto_info['precio'] * cantidad
        # Agregamos el item a la lista de compra
        items_compra.append({
            'numero': numero_producto,  # Guardamos el número
            'nombre': producto_info['nombre'],  # Guardamos el nombre
            'cantidad': cantidad,  # Guardamos la cantidad
            'precio': producto_info['precio'],  # Guardamos el precio unitario
            'total': total_item  # Guardamos el total del item
        })
        # Actualizamos el subtotal
        subtotal += total_item
        # Mostramos confirmación
        print(f"\n Agregado: {cantidad} x {producto_info['nombre']} = ${total_item:,}")
    #while para editar,eliminar y agregar productos antes de finalizar la compra
    while True:
        print("\n" + "="*70)
        print(" PRODUCTOS AGREGADOS:")
        print("="*70)
        # Verificamos si hay productos en el carrito con len de la lista items_compra
        if len(items_compra) == 0:
            print(" No hay productos en el carrito")
        else:
            # Encabezados de la tabla con formato la cual se imprimira en consola
            print(f"{'#':<5} {'PRODUCTO':<30} {'CANT.':<10} {'TOTAL':<15}")
            print("-"*70)  
            # Muestra cada producto agregado con numeracion desde 1
            for cod, item in enumerate(items_compra, 1):
                # Imprimimos con formato, utilizando f-strings para alinear columnas y utilizamos [] para acceder a los valores del diccionario
                print(f"{cod:<5} {item['nombre']:<30} {item['cantidad']:<10} ${item['total']:>13,}")
        print("="*70)

        print("\n¿Qué desea hacer?")
        print("1. Continuar con esta compra")
        print("2. Cambiar cantidad de un producto")
        print("3. Eliminar un producto")
        print("4. Agregar más productos")
        print("="*70)

        # Pedimos la opción al usuario
        opcion_edicion = input("Seleccione una opción (1-4): ")
        # Evaluamos la opción seleccionada
        if opcion_edicion == '1':
            # Si elige continuar, verificamos si hay productos en el carrito
            if len(items_compra) == 0:
                print("\n Debe agregar al menos un producto para continuar.")
                #continua el bucle principal
                continue
            # Si elige continuar, salimos del bucle de edición
            break
        # Si elige 2 sera la opcion de cambiar cantidad
        elif opcion_edicion == '2':
            # Verificamos si hay productos en el carrito
            if len(items_compra) == 0:
                print("\n No hay productos para modificar.")
                #continua el bucle principal
                continue
            #try y except para manejar errores de entrada   
            try:
                #Solicitamos el número del producto a modificar
                #(1-{len(items_compra)}) para mostrar el rango valido
                num_item = int(input(f"\n¿Qué producto desea modificar? (1-{len(items_compra)}): "))
                # Validamos que el número esté en el rango correcto
                if num_item < 1 or num_item > len(items_compra):
                    # mostramos unn mensaje de error si el numero no es valido
                    print(f"{num_item}, número inválido.")
                    #continua el bucle principal
                    continue
                # Seleccionamos el item a modificar se le resta 1 para ajustar al indice de la lista
                # Accedemos al item en la lista items_compra
                item_seleccionado = items_compra[num_item - 1]
                # Mostramos el producto y la cantidad actual
                print(f"\nProducto seleccionado: {item_seleccionado['nombre']}")
                print(f"Cantidad actual: {item_seleccionado['cantidad']}")
                # Pedimos la nueva cantidad
                nueva_cantidad = int(input("Ingrese la nueva cantidad: "))
                # Validamos que la nueva cantidad sea mayor a 0
                if nueva_cantidad <= 0:
                    print("La cantidad debe ser mayor a 0.")
                    #continua el bucle principal el cual es el while True de edicion
                    continue    
                # Actualizamos la cantidad y el total del item
                item_seleccionado['cantidad'] = nueva_cantidad
                item_seleccionado['total'] = item_seleccionado['precio'] * nueva_cantidad
                # Recalculamos el subtotal
                subtotal = sum(item['total'] for item in items_compra)
                print(f"\n Cantidad actualizada: {nueva_cantidad} x {item_seleccionado['nombre']} = ${item_seleccionado['total']:,}")
            # Manejo de errores de entrada
            except ValueError:
                print("Por favor ingrese un número válido.") 

        # Si elige 3 sera la opcion de eliminar un producto   
        elif opcion_edicion == '3':
            # Verificamos si hay productos en el carrito
            if len(items_compra) == 0:
                print("\n No hay productos para eliminar.")
                #continua el bucle principal el cual es el while True de edicion
                continue
            #try y except para manejar errores de entrada
            try:
                #Solicitamos el número del producto a eliminar
                #(1-{len(items_compra)}) para mostrar el rango valido
                num_item = int(input(f"\n¿Qué producto desea eliminar? (1-{len(items_compra)}): "))
                # Validamos que el número esté en el rango correcto
                if num_item < 1 or num_item > len(items_compra):
                    # mostramos un mensaje de error si el numero no es valido
                    print(f"{num_item}, es un número inválido.")
                    #continua el bucle principal el cual es el while True de edicion
                    continue
                # Seleccionamos el item a eliminar se le resta 1 para ajustar al indice de la lista
                item_eliminado = items_compra[num_item - 1]
                
                # Pedimos confirmación
                #.lower() para convertir la respuesta a minúsculas y facilitar la comparación
                confirmar = input(f"¿Está seguro de eliminar '{item_eliminado['nombre']}'? (si/no): ").lower()
                
                # Si confirma, eliminamos el item para eso usamos pop() para eliminar el item de la lista con el indice num_item - 1
                if confirmar == 'si':
                    items_compra.pop(num_item - 1)
                    # Recalculamos el subtotal sumando los totales de los items restantes y lo hacemos con un for 
                    subtotal = sum(item['total'] for item in items_compra)
                    # Mostramos confirmación
                    print(f"\n Producto '{item_eliminado['nombre']}' eliminado.")
                else:
                    print(f"\n  Eliminación cancelada, ingresaste 'no' o un valor incorrecto({confirmar}) .")
            # Manejo de errores de entrada
            except ValueError:
                print(" Por favor ingrese un número válido.")

        # Si elige 4 sera la opcion de agregar mas productos
        elif opcion_edicion == '4':
            # Bucle para agregar más productos
            while True:
                # Mostramos los productos disponibles
                mostrar_productos()
                #try y except para manejar errores de entrada
                try:
                    #Solicitamos el número del producto a agregar
                    numero_producto = int(input("\nIngrese el número del producto (1-30, 0 para volver): "))
                    # Validamos que el número sea válido y si es 0 para salir
                    if numero_producto == 0:
                        break  # Sale del bucle de agregar
                    # Validamos si el numero del producto esta en el diccionario productos
                    elif numero_producto in productos:
                        #si esta pedimos la cantidad
                        cantidad = int(input("¿Cuántas unidades desea? "))
                        # Validamos que la cantidad sea mayor a 0
                        if cantidad <= 0:
                            print(f"{cantidad} La cantidad debe ser mayor a 0.")
                            #continua el bucle de agregar
                            continue
                        
                        # Obtenemos la info del producto seleccionado
                        producto_info = productos[numero_producto]
                        # Calculamos el total del item y accedemos al precio del producto mediante el diccionario productos y utilizamos ['precio']
                        total_item = producto_info['precio'] * cantidad
                        
                        # Agregamos el item a la lista de compra
                        items_compra.append({
                            'numero': numero_producto,
                            'nombre': producto_info['nombre'],
                            'cantidad': cantidad,
                            'precio': producto_info['precio'],
                            'total': total_item
                        })
                        
                        # Actualizamos el subtotal
                        subtotal += total_item
                        # Mostramos confirmación
                        print(f"\n Agregado: {cantidad} x {producto_info['nombre']} = ${total_item:,}")
                        break  # Sale del bucle de agregar
                    else:
                        print(f"{numero_producto}, es un número inválido.")
                        
                except ValueError:
                    print(" Por favor ingrese un número válido.")
        
        else:
            # Si no es ninguna opción válida
            print(f"\n{opcion_edicion}, es una opción inválida. Por favor seleccione 1, 2, 3 o 4.") 


    
    # ---- CALCULAR TOTALES Y DESCUENTOS ----
    # Calculamos el subtotal sumando los totales de los items en la lista items_compra
    subtotal = sum(item['total'] for item in items_compra)
    # Si es empresa, aplicamos el descuento
    if es_empresa:
        # Calculamos el descuento y el total final con la funcion calcular_descuento
        descuento, total_final = calcular_descuento(subtotal)
    else:
        # Si no es empresa, no hay descuento
        descuento = 0
        total_final = subtotal
    # ---- MOSTRAR RESUMEN DETALLADO EN CONSOLA ----
    print("\n" + "="*70)
    print("                    RESUMEN DE LA COMPRA")
    print("="*70)
    print("FECHA Y HORA:", datos_cliente['fecha'])
    print("CLIENTE:", datos_cliente['cliente'])
    print("TIPO DE CLIENTE:", datos_cliente['tipo'])
    print("FACTURADO POR:", datos_cliente['facturador'])
    print("="*70)
    
    # Encabezados de la tabla con formato
    # Utilizamos f-strings para alinear columnas con < para izquierda y > para derecha y comas para miles
    print(f"{'PRODUCTO':<30} {'CANT.':<8} {'PRECIO UNIT.':<15} {'TOTAL':<15}")
    print("-"*70)
    
    # Mostramos cada producto
    for item in items_compra:
        print(f"{item['nombre']:<30} {item['cantidad']:<8} ${item['precio']:>13,} ${item['total']:>13,}")
    
    print("-"*70)
    
    # Si hay descuento, lo mostramos
    if es_empresa:
        print(f"{'SUBTOTAL:':<54} ${subtotal:>13,}")
        print(f"{'Descuento 5% (Empresa):':<54} -${descuento:>13,}")
        print("-"*70)
        print(f"{'TOTAL A PAGAR:':<54} ${total_final:>13,}")
    else:
        print(f"{'TOTAL A PAGAR:':<54} ${total_final:>13,}")
    
    print("="*70)

    guardar_en_excel(datos_cliente, items_compra, total_final, numero_unico_factura)

    
    while True:
        # Este bucle muestra el menú repetidamente
        
        print("\n" + "="*50)
        print("¿Qué desea hacer?")
        print("1. Imprimir factura en PDF")
        print("2. Crear nueva factura")
        print("3. Salir")
        print("="*50)
        
        # Pedimos la opción al usuario
        opcion = input("Seleccione una opción (1-3): ")
        #Dependiendo de la opción elegida se realizan diferentes acciones
        if opcion == '1':
                crear_factura_pdf(datos_cliente, items_compra, subtotal, descuento, total_final, numero_unico_factura)
                print("\n ¡Factura PDF generada exitosamente!")
        elif opcion == '2':
            
            print("\n Iniciando nueva factura...\n")
            facturacion()  
            break  
        elif opcion == '3':
            
            print("\n¡Gracias por usar nuestro sistema de facturación!")
            print("="*50)
            
            break  
            
        else:
            print(f"la opción: {opcion} es inválida. Por favor seleccione 1, 2 o 3.")
facturacion()
